from flask import Flask, jsonify, request, send_file, send_from_directory
import os
from flask_cors import CORS
from odoo_handler.models import sale_order, account_invoice
import openpyxl
from io import BytesIO
import copy
from openpyxl.styles import Font
import logging

# creating the Flask application
app = Flask(__name__, static_folder=None)
CORS(app)

# setting up logging
logging.basicConfig(filename='app.log', level=logging.DEBUG)

BUILD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '../IntraPanel/frontend/build')


@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve_react(path):
    if path and os.path.exists(os.path.join(BUILD_DIR, path)):
        return send_from_directory(BUILD_DIR, path)
    return send_from_directory(BUILD_DIR, 'index.html')


@app.route('/get-login', methods=['GET'])
def get_login():
    return {'login': 'OK', 'token': ''}


@app.route('/trex/orders')
def get_orders():
    param_search = [[['partner_id', '=', 7952], ['state', '=', 'sale'], ['invoice_status', '=', 'no']]]
    orders = sale_order.SaleOrder().get_list(param_search)
    # serializing as JSON
    return jsonify(orders)


@app.route('/trex/order_count', methods=['POST'])
def get_orders_count():
    data = request.json
    state = data.get("state")
    invoice_status = data.get("invoice_status")
    param_search = [[['partner_id', '=', 7952], ['state', '=', state], ['invoice_status', '=', invoice_status]]]
    orders = sale_order.SaleOrder().get_count(param_search)
    # serializing as JSON
    return jsonify(orders)


@app.route('/trex/invoices')
def get_invoices():
    param_search = [[['partner_id', '=', 7952], ['state', 'in', ['open']]]]
    orders = account_invoice.AccountInvoice().get_list(param_search)
    # serializing as JSON
    return jsonify(orders)


@app.route('/trex/invoice_count', methods=['POST'])
def get_invoices_count():
    data = request.json
    state = data.get("state")
    param_search = [[['partner_id', '=', 7952], ['state', '=', state]]]
    orders = account_invoice.AccountInvoice().get_count(param_search)
    # serializing as JSON
    return jsonify(orders)


# A dictionary to store files in memory
files = {}


@app.route('/upload', methods=['POST'])
def upload_file():
    filename = request.files['file'].filename
    file_to_read = request.files['file'].read()
    file_as_bytes = BytesIO(file_to_read)
    workbook = openpyxl.load_workbook(file_as_bytes, data_only=False)
    worksheet = workbook.active
    start_row = 1  # Your desired start row here
    data = []
    for row in worksheet.iter_rows(min_row=start_row, values_only=True):
        data.append([cell for cell in row])
    # Store file in files dictionary
    files[filename] = workbook
    return jsonify(data)


@app.route('/save', methods=['POST'])
def save_file():
    data = request.get_json()
    filename = data['filename']
    new_data = data['data']
    reportdata, reportdata_origin = get_eni_data_report(data['data'])

    # Get the relevant file
    workbook = files[filename]
    sheet_control = workbook.create_sheet(title="Control")
    worksheet_origin = workbook.active
    worksheet = workbook['Control']
    # cerco l'indice dove inizizare la compilazione dei campi
    initindex = None
    for index, row in enumerate(new_data):
        if len(row) > 0:
            if row[0] == 'PROTOCOLLO E PROGRESSIVO':
                initindex = index + 1
                break

    # aggiungo al file excel le celle prima dell'indice trovato
    for x, row in enumerate(reportdata[:initindex], start=1):
        for j, cell in enumerate(row, start=1):

            current_cell = worksheet_origin.cell(row=x, column=30)

            if x == initindex:
                if j == 9:
                    new_cell = worksheet.cell(row=x, column=j)
                    new_cell._style = copy.deepcopy(current_cell._style)
                    new_cell.value = "Check Transito"
                elif j == 15:
                    new_cell = worksheet.cell(row=x, column=j)
                    new_cell._style = copy.deepcopy(current_cell._style)
                    new_cell.value = "Check Unit Price"
                elif j == 25:
                    current_cell = worksheet_origin.cell(row=x, column=22)
                    new_cell = worksheet.cell(row=x, column=j)
                    new_cell._style = copy.deepcopy(current_cell._style)
                    new_cell.value = cell
                elif j == 24:
                    current_cell = worksheet_origin.cell(row=x, column=22)
                    new_cell = worksheet.cell(row=x, column=j)
                    new_cell._style = copy.deepcopy(current_cell._style)
                    new_cell.value = "Check Urgenza"
                elif j == 33:
                    new_cell = worksheet.cell(row=x, column=j)
                    new_cell._style = copy.deepcopy(current_cell._style)
                    new_cell.value = "Check Totale"
                else:
                    actual_cell = worksheet.cell(row=x, column=j)
                    actual_cell._style = copy.deepcopy(current_cell._style)
                    actual_cell.value = cell
            else:
                actual_cell = worksheet.cell(row=x, column=j)
                actual_cell._style = copy.deepcopy(current_cell._style)
                actual_cell.value = cell

        # Copy style from column 4 to column 3
        for i, cell in enumerate(worksheet['X'], start=1):
            new_cell = worksheet.cell(row=i, column=23)
            if cell.has_style:
                new_cell._style = copy.deepcopy(cell._style)

    for i, row in enumerate(reportdata[initindex:], start=initindex+1):
        note = ""
        for j, cell in enumerate(row, start=1):
            if j == 9:
                if not controls(worksheet_origin, i, cell, 'transito'):
                    add_style(worksheet, i, j)
            if j == 14:
                if not controls(worksheet_origin, i, cell, 'tariffa'):
                    add_style(worksheet, i, j)
            if j == 24:
                try:
                    val = ' / tariffa maggiorata per urgenza'
                    origin_cell = worksheet_origin.cell(row=i, column=21)
                    print(reportdata_origin[i-1][27])
                    if origin_cell.value:
                        reportdata_origin[i-1][27] = reportdata_origin[i-1][27] + val
                    if not controls(worksheet_origin, i, cell, 'urgenza'):
                        add_style(worksheet, i, j)
                        val += " ******"
                        reportdata_origin[i-1][27] = reportdata_origin[i-1][27] + val
                    note = reportdata_origin[i-1][27]
                    print(note)
                except Exception as e:
                    print(e)
            if j == 31:
                worksheet.cell(row=i, column=j, value=note)
            else:
                worksheet.cell(row=i, column=j, value=cell)

    for i, row in enumerate(reportdata_origin[initindex:], start=initindex+1):
        for j, cell in enumerate(row, start=1):
            worksheet_origin.cell(row=i, column=j, value=cell)

    # Save the updated workbook
    workbook.save(filename)
    data = []
    for row in worksheet.iter_rows(min_row=0, values_only=True):
        print(row)
        data.append([cell for cell in row])
    return jsonify(data)


def controls(sheet, row, value, type):
    var_ret = False
    if type == 'transito':
        current_cell = sheet.cell(row=row, column=8)
        if (not current_cell.value and not value) or (current_cell.value and value):
            var_ret = True

    if type == 'urgenza':
        current_cell = sheet.cell(row=row, column=21)
        if (not current_cell.value and not value) or (current_cell.value and value):
            var_ret = True

    if type == 'tariffa':
        current_cell = sheet.cell(row=row, column=13)
        if current_cell.value == value:
            var_ret = True

    if type == 'totale':
        current_cell = sheet.cell(row=row, column=29)
        if current_cell.value == value:
            var_ret = True

    return var_ret


def add_style(sheet, row, col):
    current_cell = sheet.cell(row=row, column=col)
    current_cell.font = Font(color="FF0000")


def get_eni_data_report(data):

    initindex = None
    for index, row in enumerate(data):
        if len(row) > 0:
            if row[0] == 'PROTOCOLLO E PROGRESSIVO':
                initindex = index+1
                break
    newdata = data.copy()
    newdata_origin = copy.deepcopy(data)
    # add colonne intestazione
    newdata[initindex-1].insert(8, "Check Transito")
    newdata[initindex-1].insert(14, "Check Unit Price")
    newdata[initindex-1].insert(23, "Check Urgenza")
    newdata[initindex-1].insert(32, "Check Totale")

    if initindex is not None:
        for index, row in enumerate(data[initindex:], start=initindex):
            if len(row) > 0:
                if row[0] == 'TOTALE':
                    break
                param_search = [[['service_request_customer_id', '=', row[0]]]]
                try:
                    orders = sale_order.SaleOrder().get_list(param_search)
                except Exception as e:
                    app.logger.info('get_eni_data_report')
                    app.logger.info(e)
                    app.logger.info('@@@@@@@@@@@@@@@@@@@')

                if len(orders) > 0:
                    transito = orders[0]['flex_note']
                    urgenza = orders[0]['invoice_note']
                    price_unit = next(item['price_unit'] for item in orders[0]['order_line'] if item['name'] in ('Traduzione', 'Translation'))
                    product_uom_qty = next(item['product_uom_qty'] for item in orders[0]['order_line'] if item['name'] in ('Traduzione', 'Translation'))
                    price_subtotal = next(item['price_subtotal'] for item in orders[0]['order_line'] if item['name'] in ('Traduzione', 'Translation'))

                    try:
                        ass_price_subtotal = next(item['price_subtotal'] for item in orders[0]['order_line'] if item['name'] in ('Asseverazione', 'Certification'))
                    except StopIteration:
                        ass_price_subtotal = 0

                    try:
                        lega_price_subtotal = next(item['price_subtotal'] for item in orders[0]['order_line'] if item['name'] in ('Legalizzazione', 'Legalization'))
                    except StopIteration:
                        lega_price_subtotal = 0

                    try:
                        apo_price_subtotal = next(item['price_subtotal'] for item in orders[0]['order_line'] if
                         item['name'] in ('Apostille - visto Aja', 'Apostille'))
                    except StopIteration:
                        apo_price_subtotal = 0

                    try:
                        bollo_price_subtotal = next(item['price_subtotal'] for item in orders[0]['order_line'] if
                         item['name'] in ('Imposta di bollo 16 euro\n ', '16 Euro Stamp Duty'))
                    except StopIteration:
                        bollo_price_subtotal = 0

                    try:
                        bollo_qty = next(item['product_uom_qty'] for item in orders[0]['order_line'] if
                                                    item['name'] in (
                                                    'Imposta di bollo 16 euro\n ', '16 Euro Stamp Duty'))
                    except StopIteration:
                        bollo_qty = 0

                    totale_impo_riga = price_subtotal + ass_price_subtotal + bollo_price_subtotal + lega_price_subtotal + apo_price_subtotal
                    note = str(bollo_qty)+" marche"

                    #if urgenza:
                    #    note += " / tariffa maggiorata per urgenza"
                    newdata[index].insert(8, transito)
                    newdata[index].insert(14, price_unit)
                    newdata[index][15] = product_uom_qty
                    newdata[index][16] = price_subtotal
                    newdata[index][18] = ass_price_subtotal
                    newdata[index][20] = lega_price_subtotal+apo_price_subtotal
                    newdata[index][21] = bollo_price_subtotal
                    newdata[index].insert(23, urgenza)
                    newdata[index][30] = note
                    newdata[index][31] = totale_impo_riga
                    newdata[index].insert(32, "=AF" + str(index+1) + "<>'INTRAWELT sas'!" + "AC" + str(index+1))

                    ## Importo solo i dati originali senza aggiungere colonne ##

                    newdata_origin[index][13] = product_uom_qty
                    newdata_origin[index][16] = ass_price_subtotal
                    newdata_origin[index][18] = lega_price_subtotal + apo_price_subtotal
                    newdata_origin[index][19] = bollo_price_subtotal
                    newdata_origin[index][27] = note

    return newdata, newdata_origin


@app.route('/download/<filename>', methods=['GET'])
def download_file(filename):
    # Ensure file exists.
    if not os.path.isfile(filename):
        return "File not found", 404

    return send_file(filename)


if __name__ == '__main__':
    app.run(host='127.0.0.1', port=5000)
